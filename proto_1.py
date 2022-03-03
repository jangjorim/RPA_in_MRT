
from telnetlib import SE
from tokenize import Number
import openpyxl as op

Mdl_Cd = None
SERIAL = 18401                  #   basic start serial number
IMEI = 3587770772               #   basic start IMEI number 358777077213925
COR = 0                         #   correction about serial
ImC = 11523
USIM = 898230082000498
UsC = 2111
Number_Count = 150        #   count required for work
Snc = Number_Count       #   for short
Count = 50
line = 2
i = 0

filepath = "C:/Users/MRT/python_Code/TEST_0.xlsx"

wb = op.load_workbook(filepath)
ws = wb.active






#-------------------------  Serial part -------------------------#
for i in range(SERIAL, SERIAL + Snc):
    ws.cell(row = (i - SERIAL) + 1, column = 1).value = "00" + str(i + COR)
    #print("00" + str(Snc - (Snc - i)))
    #ws.cell(row = (i - SERIAL) + 1, column = 1).value = "NULL" if (Serial_Number_Count - i) == 0 else str(Serial_Number_Count - (Serial_Number_Count - i))

#-------------------------  IMEI part   -------------------------#

for i in range(1, Snc + 1):

    ws.cell(row= i, column = 2).value = str(IMEI) + str(ImC)
    # print(i)
    # print(ImC)

    if ImC % 100 == 42:
        ImC += 17
    elif (ImC // 1000) < ((ImC + 6) // 1000):
            ImC += 6
    elif ImC % 10 == 0 :

        ImC += 10
        if (ImC // 1000) > ((ImC - 10) // 1000):
            ImC += 6
        elif (i + COR) % 10 == 3:
            ImC += 7
        else:
            ImC += 8
            
    elif ImC % 10 == 1 :

        ImC += 10

        if (ImC // 1000) > ((ImC - 10) // 1000):
            ImC += 6
        elif (i + COR) % 10 == 3:
            ImC += 7
        else:
            ImC += 8
    else :
        if (i + COR) % 10 == 3:
            ImC += 7
        else:
            ImC += 8
    
#-------------------------  USIM part   -------------------------#

for i in range(1, Snc + 1):

    ws.cell(row= i, column = 3).value = str(USIM) + str(UsC)
    print(i)
    print(UsC)

    if UsC % 100 == 42:
        UsC += 17
    elif (UsC // 1000) < ((UsC + 6) // 1000):
            UsC += 6
    elif UsC % 10 == 0:

        UsC += 10
        if (UsC // 1000) > ((UsC - 10) // 1000):
            UsC += 6
        elif (i + COR) % 10 == 4:
            UsC += 7
        else:
            UsC += 8
            
    elif UsC % 10 == 1 :

        UsC += 10

        if (UsC // 1000) > ((UsC - 10) // 1000):
            UsC += 6
        elif (i + COR) % 10 == 4:
            UsC += 7
        else:
            UsC += 8
    else :
        if (i + COR) % 10 == 4:
            UsC += 7
        else:
            UsC += 8

wb.save('Test_1.xlsx')




